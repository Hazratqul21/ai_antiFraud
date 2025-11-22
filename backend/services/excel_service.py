"""
Excel Export Service
Export transaction data and analytics to Excel files
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO
from datetime import datetime

def generate_transaction_excel(transactions: list) -> BytesIO:
    """
    Generate Excel file with transaction data
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Transactions"
    
    # Header style
    header_fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Border style
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = ['Transaction ID', 'User ID', 'Amount (so\'m)', 'Currency', 'Merchant', 
               'Location', 'IP Address', 'Device ID', 'Status', 'Timestamp', 'Risk Score']
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Data rows
    for row_num, tx in enumerate(transactions, 2):
        ws.cell(row=row_num, column=1).value = tx.transaction_id
        ws.cell(row=row_num, column=2).value = tx.user_id
        ws.cell(row=row_num, column=3).value = tx.amount
        ws.cell(row=row_num, column=4).value = tx.currency
        ws.cell(row=row_num, column=5).value = tx.merchant
        ws.cell(row=row_num, column=6).value = tx.location
        ws.cell(row=row_num, column=7).value = tx.ip_address
        ws.cell(row=row_num, column=8).value = tx.device_id
        ws.cell(row=row_num, column=9).value = tx.status
        ws.cell(row=row_num, column=10).value = tx.timestamp.strftime('%Y-%m-%d %H:%M:%S')
        
        # Risk score if available
        risk_score = tx.risk_score.score if hasattr(tx, 'risk_score') and tx.risk_score else 0
        ws.cell(row=row_num, column=11).value = risk_score
        
        # Apply borders to all cells
        for col_num in range(1, len(headers) + 1):
            ws.cell(row=row_num, column=col_num).border = thin_border
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

def generate_analytics_excel(analytics_data: dict) -> BytesIO:
    """
    Generate multi-sheet Excel with analytics data
    """
    wb = Workbook()
    
    # Summary Sheet
    ws_summary = wb.active
    ws_summary.title = "Summary"
    
    summary_data = [
        ['Metric', 'Value'],
        ['Total Transactions', analytics_data.get('total_transactions', 0)],
        ['Blocked', analytics_data.get('blocked', 0)],
        ['Under Review', analytics_data.get('challenged', 0)],
        ['Approved', analytics_data.get('allowed', 0)],
        ['Fraud Rate', f"{analytics_data.get('fraud_rate', 0):.2f}%"],
        ['Total Volume (so\'m)', analytics_data.get('total_volume', 0)],
    ]
    
    for row_num, (metric, value) in enumerate(summary_data, 1):
        ws_summary.cell(row=row_num, column=1).value = metric
        ws_summary.cell(row=row_num, column=2).value = value
    
    # Auto-adjust
    ws_summary.column_dimensions['A'].width = 25
    ws_summary.column_dimensions['B'].width = 20
    
    # By Location Sheet (if available)
    if analytics_data.get('by_location'):
        ws_location = wb.create_sheet("By Location")
        ws_location.append(['Location', 'Transaction Count', 'Fraud Rate'])
        
        for loc in analytics_data['by_location']:
            ws_location.append([
                loc['location'],
                loc['count'],
                f"{loc['fraud_rate']:.2f}%"
            ])
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
